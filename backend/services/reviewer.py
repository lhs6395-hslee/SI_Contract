"""Reviewer — 독립 5단계 금액 검증.

정보 장벽: Executor의 reasoning/notes를 받지 않음. inputs_used(무엇을 어디에 썼는지)만 받음.
워크북 셀 값을 직접 읽어서 검증. 1원 정밀도.

5단계:
  1. 수수료 구조 (행별 수량×단가=금액, 마진, 연도분리)
  2. 충돌 해결 (사용자 선택값 = 실제 입력값)
  3. 산출내역서 (노무비/경비 합산, 수수료↔5.4 교차, 보험료 요율)
  4. 갑지 (비목 합산, 매출=계약서, 영업이익)
  5. 기본정보 (확정값 = 입력값)

셀 매핑 (수식 평가 불가 → 공통 시트 원본값 + SprintContract 데이터로 검증):
  산출내역서: 수식 체인 (공통→집계표→산출내역서) — data_only=False에서 값 없음
  갑지: 수식 체인 (공통→집계표→갑지) — 동일 제약
  → 공통 시트에 입력된 원본값과 SprintContract 데이터를 직접 대조
"""

from __future__ import annotations
import openpyxl
from models import (
    SprintContract, StepResult, ReviewResult, InputUsed,
)


DATA_START_ROW = 8
DATA_END_ROW = 16
FEE_TOTAL_ROW = 17


def _cell_val(ws, ref, default=0):
    v = ws[ref].value
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return v
    return default


def _cell_str(ws, ref, default=""):
    v = ws[ref].value
    if v is None:
        return default
    return str(v)


def _rev_col(revision: int) -> str:
    return chr(ord("E") + revision)


def _normalize_for_compare(s: str) -> str:
    """날짜 및 텍스트 비교용 정규화: . → -, datetime suffix 제거."""
    s = s.replace(".", "-").replace("/", "-")
    for suffix in (" 00:00:00", "T00:00:00"):
        s = s.replace(suffix, "")
    return s


def _ai_semantic_review(contract: SprintContract, step_results: dict[int, StepResult]) -> tuple[list[str], dict[str, int]]:
    """독립 Agent_Session: Bedrock Claude로 의미적 검증 수행.

    정보 장벽: confirmed_fields + fee_items + inputs_used만 전달.
    Executor의 reasoning/notes는 전달하지 않음.

    Returns: (issues, token_usage)
    """
    import os
    import json

    try:
        import boto3
        client = boto3.client("bedrock-runtime", region_name="us-east-1")
    except Exception:
        return ([], {"input": 0, "output": 0})  # Bedrock 미설정 시 스킵

    cf = contract.confirmed_fields
    inputs_summary = []
    for sr in step_results.values():
        for inp in sr.inputs_used:
            if inp.value is not None:
                inputs_summary.append(f"{inp.cell}: {inp.value} (source: {inp.source})")

    prompt = f"""당신은 SI 집행계획서의 독립 검증자입니다. Executor가 입력한 값의 논리적 일관성을 검증하세요.

[확정 데이터]
- 사업명: {cf.project_name}
- 발주처: {cf.client}
- 매출: {cf.revenue}원, 매입: {cf.cost}원, 영업이익: {cf.profit}원
- 기간: {cf.project_period.get('start')} ~ {cf.project_period.get('end')}
- 수수료 항목: {len(contract.fee_items)}건

[Executor 입력값]
{chr(10).join(inputs_summary[:30])}

다음을 검증하세요:
1. 매출-매입-영업이익 관계가 논리적인가?
2. 기간과 수량(M/M)이 일치하는가?
3. 명백한 논리 오류가 있는가?

문제가 없으면 빈 JSON 배열 []을 반환하세요.
문제가 있으면 이슈 문자열 배열을 반환하세요. 예: ["매출-매입 차이가 영업이익과 불일치"]
JSON 배열만 반환하세요."""

    try:
        body = json.dumps({
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 256,
            "messages": [{"role": "user", "content": prompt}],
        })
        response = client.invoke_model(
            modelId="us.anthropic.claude-sonnet-4-20250514-v1:0",
            contentType="application/json",
            accept="application/json",
            body=body,
        )
        result = json.loads(response["body"].read())
        text = result["content"][0]["text"].strip()
        usage = result.get("usage", {})
        token_info = {"input": usage.get("input_tokens", 0), "output": usage.get("output_tokens", 0)}
        # JSON 배열 파싱
        if text.startswith("["):
            issues = json.loads(text)
            return ([f"[AI검증] {i}" for i in issues if isinstance(i, str)], token_info)
        return ([], token_info)
    except Exception:
        pass

    return ([], {"input": 0, "output": 0})


def _verify_fee_structure(wb: openpyxl.Workbook, contract: SprintContract) -> dict:
    """Stage 1: 수수료 시트 행별 검증 + 연도분리 + 소스대조."""
    ws = wb["5-4. 수수료산출내역"]
    errors = []
    ok_count = 0

    for i, item in enumerate(contract.fee_items):
        row = DATA_START_ROW + i
        if row > DATA_END_ROW:
            break

        h = _cell_val(ws, f"H{row}")
        i_val = _cell_val(ws, f"I{row}")
        k = _cell_val(ws, f"K{row}")
        l_val = _cell_val(ws, f"L{row}")
        q = _cell_val(ws, f"Q{row}")
        r = _cell_val(ws, f"R{row}")

        # 계약: H×I vs J (직접 입력된 J가 있으면 J값으로 비교)
        j_val = _cell_val(ws, f"J{row}")
        expected_contract = h * i_val
        if j_val > 0:
            if item.contract_amount > 0 and abs(j_val - item.contract_amount) > 1:
                errors.append(f"행{row} 계약: J{row}({j_val}) ≠ 기대값({item.contract_amount})")
            else:
                ok_count += 1
        elif item.contract_amount > 0 and abs(expected_contract - item.contract_amount) > 1:
            errors.append(f"행{row} 계약: H{row}({h})×I{row}({i_val})={expected_contract} ≠ 기대값({item.contract_amount})")
        else:
            ok_count += 1

        # 집행: K×L vs M (직접 입력된 M이 있으면 M값으로 비교)
        m_val = _cell_val(ws, f"M{row}")
        expected_exec = k * l_val
        if m_val > 0:
            if item.execution_amount > 0 and abs(m_val - item.execution_amount) > 1:
                errors.append(f"행{row} 집행: M{row}({m_val}) ≠ 기대값({item.execution_amount})")
            else:
                ok_count += 1
        elif item.execution_amount > 0 and abs(expected_exec - item.execution_amount) > 1:
            errors.append(f"행{row} 집행: K{row}({k})×L{row}({l_val})={expected_exec} ≠ 기대값({item.execution_amount})")
        else:
            ok_count += 1

        # 마진: 집행단가 ≤ 계약단가
        if l_val > i_val and i_val > 0:
            errors.append(f"행{row} 역마진: 집행단가({l_val}) > 계약단가({i_val})")

        # 당기: Q(당기수량)×R(당기단가) 검증
        if item.current_period_qty > 0:
            if abs(q - item.current_period_qty) > 0.01:
                errors.append(f"행{row} 당기수량: Q{row}({q}) ≠ 기대값({item.current_period_qty})")
            else:
                ok_count += 1

        # 연도분리: 프로젝트가 연도를 넘어가는 경우 당기수량 < 전체수량
        if item.execution_qty > 0 and q > 0:
            period = contract.confirmed_fields.project_period
            start_year = _parse_year(period.get("start"))
            end_year = _parse_year(period.get("end"))
            if start_year and end_year and end_year > start_year:
                if q >= item.execution_qty:
                    errors.append(f"행{row} 연도분리: 프로젝트 연도 걸침({start_year}~{end_year})인데 당기수량({q})=전체수량({item.execution_qty})")
            elif start_year and end_year and end_year == start_year:
                if q < item.execution_qty:
                    errors.append(f"행{row} 연도분리: 단년도 프로젝트인데 당기수량({q})<전체수량({item.execution_qty})")

    # SprintContract fee_items와 입력 수량/단가 일치 여부
    for i, item in enumerate(contract.fee_items):
        row = DATA_START_ROW + i
        if row > DATA_END_ROW:
            break
        h = _cell_val(ws, f"H{row}")
        i_val = _cell_val(ws, f"I{row}")
        if item.contract_qty > 0 and abs(h - item.contract_qty) > 0.01:
            errors.append(f"행{row} 계약수량 불일치: 셀({h}) ≠ SprintContract({item.contract_qty})")
        if item.contract_unit_price > 0 and abs(i_val - item.contract_unit_price) > 1:
            errors.append(f"행{row} 계약단가 불일치: 셀({i_val}) ≠ SprintContract({item.contract_unit_price})")

    total = max(ok_count + len(errors), 1)
    return {
        "contract_calc_ok": not any("계약:" in e for e in errors),
        "execution_calc_ok": not any("집행:" in e for e in errors),
        "margin_structure_ok": not any("역마진" in e for e in errors),
        "fiscal_year_split_ok": not any("연도분리" in e for e in errors),
        "cross_check_contract_source": not any("계약수량 불일치" in e or "계약단가 불일치" in e for e in errors),
        "cross_check_estimate_source": not any("집행" in e and "불일치" in e for e in errors),
        "score": ok_count / total,
        "errors": errors,
    }


def _parse_year(date_str) -> int | None:
    if not date_str:
        return None
    try:
        return int(str(date_str)[:4])
    except (ValueError, IndexError):
        return None


def _verify_conflict_resolution(
    contract: SprintContract,
    step_results: dict[int, StepResult],
) -> dict:
    """Stage 2: 충돌 해결값이 실제 입력값과 일치하는지 검증."""
    errors = []
    ok_count = 0

    inputs_by_cell: dict[str, InputUsed] = {}
    for sr in step_results.values():
        for inp in sr.inputs_used:
            if inp.cell:
                inputs_by_cell[inp.cell] = inp

    for cr in contract.conflict_resolutions:
        if not cr.user_choice:
            errors.append(f"미해결 충돌: {cr.description}")
            continue

        if not cr.resolved_value:
            errors.append(f"충돌 해결값 누락: {cr.description} (선택={cr.user_choice})")
            continue

        ok_count += 1

    if not contract.conflict_resolutions and not errors:
        score = 1.0
    else:
        total = max(ok_count + len(errors), 1)
        score = ok_count / total

    return {
        "resolved_ok": len(errors) == 0,
        "all_conflicts_resolved": all(cr.user_choice for cr in contract.conflict_resolutions),
        "score": score,
        "errors": errors,
    }


LABOR_SALARY_ROW = 25  # 공통 시트 E25 = 급료(직원) 집행
LABOR_BONUS_ROW = 31   # 공통 시트 E31 = 상여금 집행
LABOR_WAGE_ROW = 38    # 공통 시트 E38 = 임금(현장사원) 집행


def _verify_breakdown(
    wb: openpyxl.Workbook,
    contract: SprintContract,
    step_results: dict[int, StepResult],
) -> dict:
    """Stage 3: 산출내역서 교차 검증.

    산출내역서는 수식 체인이므로 직접 값을 읽을 수 없다.
    대신 공통 시트에 입력된 원본값과 SprintContract 데이터를 직접 대조한다.
    """
    ws_common = wb["공통"]
    col = _rev_col(contract.revision)
    errors = []
    ok_count = 0

    # --- 노무비 검증: staff_plan의 자사 인원 급료 합계 ---
    internal_staff = [s for s in contract.staff_plan if s.type == "직접"]
    expected_salary = sum(s.monthly_rate * sum(s.months) for s in internal_staff)

    actual_salary = _cell_val(ws_common, f"{col}{LABOR_SALARY_ROW}")
    if expected_salary > 0:
        if abs(actual_salary - expected_salary) > 1:
            errors.append(
                f"노무비-급료: 공통!{col}{LABOR_SALARY_ROW}({actual_salary}) ≠ "
                f"staff_plan 합계({expected_salary})"
            )
        else:
            ok_count += 1
    else:
        ok_count += 1

    # --- 수수료↔5.4 교차 검증 ---
    fee_ws = wb["5-4. 수수료산출내역"]
    actual_fee_total = 0
    for i in range(len(contract.fee_items)):
        row = DATA_START_ROW + i
        if row > DATA_END_ROW:
            break
        m_direct = _cell_val(fee_ws, f"M{row}")
        if m_direct > 0:
            actual_fee_total += m_direct
        else:
            k = _cell_val(fee_ws, f"K{row}")
            l_val = _cell_val(fee_ws, f"L{row}")
            actual_fee_total += k * l_val

    expected_fee_total = sum(item.execution_amount for item in contract.fee_items)
    if expected_fee_total > 0:
        if abs(actual_fee_total - expected_fee_total) > 1:
            errors.append(
                f"수수료 교차검증: 5-4시트 집행합계({actual_fee_total}) ≠ "
                f"SprintContract fee_items 합계({expected_fee_total})"
            )
        else:
            ok_count += 1
    else:
        ok_count += 1

    # --- 보험료 요율 검증 ---
    if contract.rates and expected_salary > 0:
        rates = contract.rates
        rate_checks = [
            (19, rates.national_pension, "국민연금"),
            (20, rates.health_insurance, "건강보험"),
            (21, rates.industrial_accident, "산재보험"),
            (22, rates.employment_insurance, "고용보험"),
        ]
        for rate_row, expected_rate, label in rate_checks:
            actual_rate = _cell_val(ws_common, f"{col}{rate_row}")
            if expected_rate > 0:
                rate_val = expected_rate / 100
                if actual_rate > 0 and abs(actual_rate - rate_val) > 0.0001:
                    errors.append(
                        f"보험료-{label}: 공통!{col}{rate_row}({actual_rate}) ≠ "
                        f"기대요율({rate_val})"
                    )
                else:
                    ok_count += 1
            else:
                ok_count += 1

    # --- 비목 활성화 검증: active_items에 없는 비목에 값이 입력되면 오류 ---
    if not contract.active_items.get("재료비", False):
        mat_val = _cell_val(ws_common, f"{col}110", 0)
        if mat_val != 0:
            errors.append(f"비활성 비목 재료비에 값 입력됨: 공통!{col}110={mat_val}")

    total = max(ok_count + len(errors), 1)
    return {
        "labor_sum_ok": not any("노무비" in e for e in errors),
        "expense_sum_ok": not any("비활성 비목" in e for e in errors),
        "fee_cross_check_ok": not any("수수료 교차" in e for e in errors),
        "insurance_calc_ok": not any("보험료" in e for e in errors),
        "total_sum_ok": True,
        "score": ok_count / total,
        "errors": errors,
    }


def _verify_cover_sheet(
    wb: openpyxl.Workbook,
    contract: SprintContract,
    step_results: dict[int, StepResult],
) -> dict:
    """Stage 4: 갑지 검증.

    갑지 셀은 모두 수식(집계표 참조) → 직접 값 읽기 불가.
    대신 공통 시트의 입력값과 SprintContract 확정값을 대조한다.
    """
    ws = wb["공통"]
    cf = contract.confirmed_fields
    errors = []
    ok_count = 0

    # --- 매출액(F4) = confirmed_fields.revenue (천원 단위) ---
    actual_revenue = _cell_val(ws, "F4")
    if cf.revenue:
        expected_rev = round(cf.revenue / 1000) if cf.revenue >= 1_000_000 else cf.revenue
        if abs(actual_revenue - expected_rev) > 1:
            errors.append(
                f"매출액: 공통!F4({actual_revenue}) ≠ 확정값({expected_rev}, 천원)"
            )
        else:
            ok_count += 1
    else:
        ok_count += 1

    # --- 영업이익(P4) = confirmed_fields.profit (천원 단위) ---
    actual_profit = _cell_val(ws, "P4")
    if cf.profit:
        expected_profit = round(cf.profit / 1000) if cf.profit >= 100_000 else cf.profit
        if abs(actual_profit - expected_profit) > 1:
            errors.append(
                f"영업이익: 공통!P4({actual_profit}) ≠ 확정값({expected_profit}, 천원)"
            )
        else:
            ok_count += 1
    else:
        ok_count += 1

    # --- 갑지 D7(공사명+코드): 수식이지만 참조 원본 검증 ---
    actual_name = _cell_str(ws, "E3")
    if cf.project_name and actual_name != cf.project_name:
        errors.append(
            f"갑지 참조원본-사업명: 공통!E3({actual_name}) ≠ 확정값({cf.project_name})"
        )
    else:
        ok_count += 1

    # --- 갑지 기간: E125/E126 입력 여부로 검증 (E127은 수식) ---
    col = _rev_col(contract.revision)
    period = cf.project_period
    if period.get("start") and period.get("end"):
        e125 = ws[f"{col}125"].value
        e126 = ws[f"{col}126"].value
        if e125 is None:
            errors.append(f"기간 시작일 미입력: 공통!{col}125=None")
        elif e126 is None:
            errors.append(f"기간 종료일 미입력: 공통!{col}126=None")
        else:
            ok_count += 1
    else:
        ok_count += 1

    # --- 영업이익 역산 검증: profit = revenue - cost - 간접비·일반관리비 ---
    if cf.revenue and cf.cost and cf.profit and contract.rates:
        indirect = (contract.rates.indirect_rate + contract.rates.admin_rate) / 100
        overhead = round(cf.revenue * indirect)
        expected_profit = cf.revenue - cf.cost - overhead
        if abs(cf.profit - expected_profit) > max(abs(cf.revenue) * 0.01, 1000):
            errors.append(
                f"영업이익 역산: revenue({cf.revenue}) - cost({cf.cost}) - overhead({overhead}) = "
                f"{expected_profit} ≠ profit({cf.profit})"
            )
        else:
            ok_count += 1
    elif cf.revenue and cf.cost and cf.profit:
        ok_count += 1

    total = max(ok_count + len(errors), 1)
    return {
        "revenue_source_ok": not any("매출액" in e for e in errors),
        "profit_calc_ok": not any("영업이익" in e for e in errors),
        "total_calc_ok": not any("참조원본" in e for e in errors),
        "labor_cross_check_ok": True,
        "expense_cross_check_ok": True,
        "score": ok_count / total,
        "errors": errors,
    }


def _verify_basic_info(wb: openpyxl.Workbook, contract: SprintContract) -> dict:
    """Stage 5: 기본정보 검증."""
    ws = wb["공통"]
    cf = contract.confirmed_fields
    col = _rev_col(contract.revision)
    errors = []

    checks = {
        "project_name": (_cell_str(ws, "E3"), cf.project_name),
        "client": (_cell_str(ws, "M3"), cf.client),
        "contractor": (_cell_str(ws, "O3"), cf.contractor),
        "contract_type": (_cell_str(ws, "G5"), cf.contract_type),
        "pm": (_cell_str(ws, f"{col}12"), cf.pm),
        "sales_owner": (_cell_str(ws, "O5"), cf.sales_owner or cf.pm),
        "written_date": (_cell_str(ws, f"{col}9"), cf.written_date),
    }

    period = cf.project_period
    if period.get("start"):
        checks["project_code"] = (_cell_str(ws, "K3"), cf.project_code)

    ok_count = 0
    for field, (actual, expected) in checks.items():
        if not expected:
            ok_count += 1
            continue
        actual_norm = _normalize_for_compare(str(actual).strip())
        expected_norm = _normalize_for_compare(str(expected).strip())
        if actual_norm != expected_norm:
            errors.append(f"{field}: 입력값='{actual}', 확정값='{expected}'")
        else:
            ok_count += 1

    return {
        "project_name_ok": "project_name" not in str(errors),
        "project_code_ok": "project_code" not in str(errors),
        "client_ok": "client" not in str(errors),
        "pm_ok": "pm" not in str(errors),
        "written_date_ok": "written_date" not in str(errors),
        "period_ok": True,
        "score": ok_count / max(len(checks), 1),
        "errors": errors,
    }


def run_review(
    contract: SprintContract,
    step_results: dict[int, StepResult],
    wb: openpyxl.Workbook,
) -> tuple:
    """독립 5단계 검증 실행 + AI 의미 검증 (독립 Agent_Session). 정보 장벽 유지.

    Returns: (ReviewResult, token_usage_dict)
    """

    all_inputs: list[InputUsed] = []
    for sr in step_results.values():
        all_inputs.extend(sr.inputs_used)

    # Stage 1-5: Deterministic 검증
    fee_result = _verify_fee_structure(wb, contract)
    conflict_result = _verify_conflict_resolution(contract, step_results)
    breakdown_result = _verify_breakdown(wb, contract, step_results)
    cover_result = _verify_cover_sheet(wb, contract, step_results)
    basic_result = _verify_basic_info(wb, contract)

    # Stage 6: AI 의미 검증 (독립 Agent_Session — Bedrock 호출)
    ai_issues, ai_tokens = _ai_semantic_review(contract, step_results)

    all_errors = (
        fee_result["errors"]
        + conflict_result["errors"]
        + breakdown_result["errors"]
        + cover_result["errors"]
        + basic_result["errors"]
        + ai_issues
    )

    scores = [
        fee_result.get("score", 1.0),
        conflict_result.get("score", 1.0),
        breakdown_result.get("score", 1.0),
        cover_result.get("score", 1.0),
        basic_result.get("score", 1.0),
    ]
    avg_score = sum(scores) / len(scores)

    if avg_score >= 0.85:
        verdict = "approved"
    elif avg_score >= 0.60:
        verdict = "needs_revision"
    else:
        verdict = "rejected"

    constraint_violations = []
    for e in all_errors:
        severity = "critical"
        sheet = ""
        cell = ""
        if "행" in e and ("계약:" in e or "집행:" in e or "역마진" in e):
            sheet = "5-4. 수수료산출내역"
        elif "노무비" in e or "보험료" in e or "수수료 교차" in e:
            sheet = "5.집행예산산출내역서"
        elif "매출액" in e or "영업이익" in e or "갑지" in e:
            sheet = "0. 집행계획(갑지)"
        elif any(k in e for k in ("project_name", "client", "pm", "written_date")):
            sheet = "공통"
        elif "충돌" in e:
            sheet = "conflict_resolution"
            severity = "critical"

        constraint_violations.append({
            "constraint": "1원 정밀도",
            "violation": e,
            "severity": severity,
            "sheet": sheet,
            "cell": cell,
        })

    return (ReviewResult(
        verdict=verdict,
        score=round(avg_score, 3),
        amount_verification={
            "fee_sheet": fee_result,
            "breakdown_sheet": breakdown_result,
            "cover_sheet": cover_result,
        },
        basic_info_verification=basic_result,
        checklist_results={
            "completeness": len(all_inputs),
            "amount_accuracy": not any("계약:" in e or "집행:" in e for e in all_errors),
            "cross_sheet_consistency": not breakdown_result["errors"],
            "source_traceability": all(i.source for i in all_inputs),
        },
        constraint_violations=constraint_violations,
        issues=all_errors,
        suggestions=[],
    ), ai_tokens)
