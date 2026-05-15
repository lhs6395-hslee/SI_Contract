"""3. 예정공정표 시트 라이터.

공종: C열 (11행부터)
공사기간: H~AF열 (월별 셀 병합/색상으로 간트 표시)
"""

from .base import SheetWriter
from openpyxl.styles import PatternFill

GANTT_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DATA_START_ROW = 11


class ScheduleSheetWriter(SheetWriter):
    sheet_name = "3. 예정공정표"

    def _write(self):
        items = self.contract.schedule
        if not items:
            return

        for i, item in enumerate(items):
            row = DATA_START_ROW + i
            self.write_cell(f"C{row}", item.name, source=f"schedule[{i}].name")

            for m in range(item.start_month, item.end_month + 1):
                col = chr(ord("H") + m) if m < 18 else chr(ord("A")) + chr(ord("A") + m - 18)
                try:
                    cell = self.ws[f"{col}{row}"]
                    cell.fill = GANTT_FILL
                except Exception:
                    pass
