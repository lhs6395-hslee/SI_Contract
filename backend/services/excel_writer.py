"""집행계획서 엑셀 생성 — 템플릿 기반 값 삽입"""

import os
import tempfile
import unicodedata
from pathlib import Path
from copy import copy
from openpyxl import load_workbook


def resolve_template_path(templates_dir: Path, filename: str = "템플릿.xlsx") -> Path:
    """한글 파일명 NFC/NFD 정규화 차이에 안전한 템플릿 경로 해석.

    macOS에서 docker build 시 파일명이 NFD 바이트로 이미지에 들어가
    Linux에서 NFC 경로 조회가 실패한다 — 디렉토리를 스캔해 정규화 일치로 찾는다.
    """
    direct = templates_dir / filename
    if direct.exists():
        return direct
    if templates_dir.is_dir():
        target = unicodedata.normalize("NFC", filename)
        for name in os.listdir(templates_dir):
            if unicodedata.normalize("NFC", name) == target:
                return templates_dir / name
    return direct  # 미존재 — 호출부에서 FileNotFoundError 처리


# 템플릿 경로 (프로젝트 루트 기준)
TEMPLATE_PATH = resolve_template_path(Path(__file__).parent.parent.parent / "templates")


def generate_excel(data: dict) -> str:
    """추출/수정된 데이터로 집행계획서 엑셀을 생성하고 파일 경로를 반환한다.

    data 구조 예시:
    {
      "projectName": "퀘이사존 유지보수",
      "client": "삼성SDS",
      "contractor": "GS네오텍",
      "startDate": "2026.01.01",
      "endDate": "2026.12.31",
      "revenue": 156600000,
      "cost": 98400000,
      ...
    }
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"템플릿 파일이 없습니다: {TEMPLATE_PATH}")

    wb = load_workbook(str(TEMPLATE_PATH))

    # ── 공통시트 E열 채우기 ──
    if "공통" in wb.sheetnames:
        ws = wb["공통"]
        _fill_common_sheet(ws, data)

    # ── 기타 시트 직접 입력 셀 (향후 확장) ──
    # 5-4.수수료산출내역 등은 별도 데이터가 오면 채움

    # ── 임시 파일로 저장 ──
    out = Path(tempfile.mkdtemp()) / "집행계획서.xlsx"
    wb.save(str(out))
    wb.close()
    return str(out)


# ─── 공통시트 매핑 ─────────────────────────────────────────

# 공통시트 E열 셀 매핑 (행번호 → data 키)
COMMON_MAPPING = {
    9: "projectName",      # 사업명
    10: "client",           # 발주처
    11: "contractor",       # 계약처
    14: "contractType",     # 계약구분
    18: "startDate",        # 계약시작
    19: "endDate",          # 계약종료
    22: "pm",               # PM
    23: "salesOwner",       # 영업담당
    28: "paymentTerms",     # 수금조건
    # 금액 관련
    135: "revenue",         # 매출합계
    136: "cost",            # 매입합계
}


def _fill_common_sheet(ws, data: dict):
    """공통시트 E열에 값 삽입."""
    for row, key in COMMON_MAPPING.items():
        val = data.get(key)
        if val is not None:
            # 딕셔너리 형태({value: ...})이면 value 추출
            if isinstance(val, dict):
                val = val.get("value")
            if val is not None:
                ws.cell(row=row, column=5, value=val)  # E열 = column 5
