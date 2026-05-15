"""
집행계획서 템플릿 생성 스크립트 v5
AI가 직접 값을 입력해야 하는 셀      → 연한 초록색  (CCFFCC)
AI가 수식을 작성해야 하는 셀          → 연한 노란색  (FFFFCC)
소스 없이 수동 입력이 필요한 셀       → 연한 주황색  (FFD9B3)
수식/자동계산/레이블/헤더 셀          → 변경 없음

선정 기준:
- 초록: 소스 자료(계약서/견적서)에서 값을 추출하거나 계산해서 채워야 하는 셀
- 노란: 현재 하드코딩이지만 수식으로 대체해야 하는 셀
- 주황: 소스 자료에 없고 PM/담당자가 경험적으로 판단해서 입력하는 셀
"""

import shutil
import openpyxl
from openpyxl.styles import PatternFill

SRC = "origin/1. (최초) 집행계획서_25년 삼성전자DS MSP_v0.3_FN.xlsx"
DST = "origin/집행계획서_템플릿.xlsx"
GREEN  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
ORANGE = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")
GREEN_RGB_VARIANTS  = {"CCFFCC",  "00CCFFCC"}
YELLOW_RGB_VARIANTS = {"FFFFCC",  "00FFFFCC"}
ORANGE_RGB_VARIANTS = {"FFD9B3",  "00FFD9B3"}
ALL_MARKED = GREEN_RGB_VARIANTS | YELLOW_RGB_VARIANTS | ORANGE_RGB_VARIANTS

def mark(ws, cell_ref):
    ws[cell_ref].fill = GREEN

def mark_yellow(ws, cell_ref):
    ws[cell_ref].fill = YELLOW

def mark_orange(ws, cell_ref):
    ws[cell_ref].fill = ORANGE

def mark_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.fill = GREEN

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)

# ══════════════════════════════════════════════════════
# 공통2 시트
# ══════════════════════════════════════════════════════
ws = wb["공통2"]

# 집행 차수 (본 집행=0, 수정=1~)
mark(ws, "D4")
# 프로젝트 기본 정보
for c in ["D5", "I5", "D6", "D7", "D8", "D9", "D10"]:
    mark(ws, c)
# 고객사 담당자
for c in ["F7", "H7"]:
    mark(ws, c)
# 고객사 주소
mark(ws, "H6")
# 투입인원: 역할명(D) / 이름(E) / 직급(F) / MM(G) — 행16~19
mark_range(ws, 16, 19, 4, 7)   # D~G열
# 작성일, 시작일, 종료일
for c in ["E23", "E24", "E25"]:
    mark(ws, c)
# 견적 제출일, 예정가격
for c in ["E28", "E29"]:
    mark(ws, c)
# 견적금액 비목별 — E30(재료비)은 수식참조, E35(재료비계약)도 수식참조 → 제외
# 노무비/외주비/경비만 AI입력
for c in ["E31", "E32", "E33"]:
    mark(ws, c)
# 집행금액 섹션 외주비 (F37 = 0, 값 셀)
mark(ws, "F37")
# 매출금액
mark(ws, "F42")
# 계약금액 합계 및 비목별 (숨겨진 참조값)
for c in ["E55", "E57", "E58", "E59"]:
    mark(ws, c)
# E30(재료비)=수식참조, E34(합계)=수식 → 제외. 노무비/외주비/경비만 추가 표시
for c in ["E31", "E32", "E33"]:
    mark(ws, c)
# 수주손익표 매출액 비율 행 (행67: 고정 100%)
for col in [5, 7, 9, 11, 15, 17, 19, 21, 23, 25]:  # E,G,I,K,O,Q,S,U,W,Y
    ws.cell(row=67, column=col).fill = GREEN
# 수주손익표 재료비 행 (행68) F열
ws.cell(row=68, column=6).fill = GREEN

# ══════════════════════════════════════════════════════
# 5.4 수수료 시트
# ══════════════════════════════════════════════════════
ws = wb["5.4 수수료"]

# 데이터 행: 6~8(용역), 10~12(유지보수), 14~16(기타)
data_rows = list(range(6, 9)) + list(range(10, 13)) + list(range(14, 17))

# AI 입력 열: B(코드)=2, C(품명)=3, D(규격)=4, E(단위)=5,
#             F(계약수량)=6, G(계약단가)=7,
#             I(집행단위)=9, J(집행수량)=10, K(집행단가)=11,
#             M(당기수량)=13, S(비고)=19
# 수식 열: H=8, L=12, N=14, O=15, P=16, Q=17, R=18 → 제외
ai_cols = [2, 3, 4, 5, 6, 7, 9, 10, 11, 13, 19]

for r in data_rows:
    for c in ai_cols:
        ws.cell(row=r, column=c).fill = GREEN

# ══════════════════════════════════════════════════════
# 5. 산출내역서 시트
# ══════════════════════════════════════════════════════
ws = wb["5. 산출내역서"]

# ── 재료비 행8 (현지재료비) — E/F/G 값 셀 ──
for col in [5, 6, 7]:   # E,F,G
    ws.cell(row=8, column=col).fill = GREEN

# ── D열(산출근거 텍스트) ──
# AI가 소스 자료 기반으로 산출근거를 기술하는 행
# D26~D30(보험료 노무): 원본에 이미 요율 설명 텍스트가 고정값으로 있음 → 제외
d_rows = [10, 12, 14, 15, 17,       # 노무비 (급료/임금/상여/퇴직)
          23,                         # 복리후생비
          32, 33, 34,                 # 보험료(공사) — 보증 종류
          42, 48,                     # 여비교통비, 차량유지비
          50, 51, 52]                 # 수수료
for r in d_rows:
    ws.cell(row=r, column=4).fill = GREEN

# ── E열(계약금액) ── 수식 셀 제외
# 수식: 6,7,9,11,16,19,20,21,22,24,25,31,38,39,40,41,49,50,51,52,54
# 값 셀 (0 포함 — 프로젝트마다 달라짐):
e_rows = [8,                          # 현지재료비
          10, 12, 13, 14, 15, 17, 18, # 노무비
          23,                          # 복리후생비
          26, 27, 28, 29, 30,          # 보험료(노무) — 계약금액은 별도 입력
          32, 33, 34,                  # 보험료(공사)
          35, 36, 37,                  # 임차료/운반/수선
          42, 43, 44, 45, 46, 47,      # 여비/통신/접대/도서/교육/안전
          53]                          # 감가상각비
for r in e_rows:
    ws.cell(row=r, column=5).fill = GREEN

# ── F열(집행금액) ── 수식 셀 제외
# 수식: 6,7,9,10,11,14,15,19,20,21,22,23,24,25,26,27,29,30,31,38,39,40,41,42,48,49,50,51,54
# 소모품비(38~41): E/F 모두 SUMIF 수식 → 값 셀 없음 (5.3 소모품비 시트에서 입력)
# 값 셀:
f_rows = [8,                          # 현지재료비
          12, 13, 16, 17, 18,          # 현장임금/퇴직금현장
          28,                          # 퇴직공제부금(=0 고정)
          32, 33, 34,                  # 보험료(공사)
          35, 36, 37,                  # 임차료/운반/수선
          43, 44, 45, 46, 47,          # 통신/접대/도서/교육/안전
          52,                          # 기타수수료
          53]                          # 감가상각비
for r in f_rows:
    ws.cell(row=r, column=6).fill = GREEN

# ── G열(정산금액) ── 합계수식 행 제외
# 수식: 6,9,11,16(SUM),19,22,24,25,31,38,49,54
# 값 셀 (대부분 0 — 정산 시 채움):
g_rows = [7, 8,                        # 재료비
          10, 12, 13, 14, 15, 17, 18,  # 노무비
          20, 21,                       # 외주비
          23,                           # 복리후생비
          26, 27, 28, 29, 30,           # 보험료(노무)
          32, 33, 34,                   # 보험료(공사)
          35, 36, 37,                   # 임차료/운반/수선
          39, 40, 41,                   # 소모품비 정산 (G는 값 셀)
          42, 43, 44, 45, 46, 47,       # 여비/통신/접대/도서/교육/안전
          50, 51, 52,                   # 수수료
          53]                           # 감가상각비
for r in g_rows:
    ws.cell(row=r, column=7).fill = GREEN

# ══════════════════════════════════════════════════════
# 5.3 소모품비 시트
# ══════════════════════════════════════════════════════
ws = wb["5.3 소모품비"]

# 행6: 데이터 행 (행7은 합계 수식)
# B(코드)=2, C(품명)=3, D(규격)=4, E(단위)=5
# F(계약수량)=6, G(계약단가)=7 — H(계약금액)=8 수식
# I(집행수량)=9, J(집행단가)=10 — K(집행금액)=11 수식
# L(당기수량)=12, M(당기금액)=13
# N(차기수량)=14 — O(차기금액)=15 수식
# 수식 열: H=8, K=11, O=15, P=16, Q=17 → 제외
소모품_ai_cols = [2, 3, 4, 5, 6, 7, 9, 10, 12, 13, 14]
for c in 소모품_ai_cols:
    ws.cell(row=6, column=c).fill = GREEN

# 5.4 수수료 O열(차기이월수량) — 행10은 값 셀(수식 없음), 나머지는 수식
ws = wb["5.4 수수료"]
ws.cell(row=10, column=15).fill = GREEN  # O10

# ══════════════════════════════════════════════════════
# 0. 집행(갑지) 시트
# ══════════════════════════════════════════════════════
ws = wb["0. 집행(갑지)"]

# 수주손익 금액 + 비율
for c in ["F13", "F15", "F17", "F19"]:
    mark(ws, c)
for c in ["G13", "G15", "G17", "G18"]:
    mark(ws, c)
# 계약금액 비율 — I13, K13은 AI 입력값(100 고정)
for c in ["I13", "K13"]:
    mark(ws, c)
# S13, U13: 행14~19와 동일한 수식 패턴으로 변환 (노란색)
# 행13이 기준행이므로 수식 결과는 항상 100이지만 수식으로 일관성 유지
ws["S13"].value = "=IFERROR((R13/R$13*100),0)"
mark_yellow(ws, "S13")
ws["U13"].value = "=IFERROR((T13/T$13*100),0)"
mark_yellow(ws, "U13")
# 정산누계 매출액
mark(ws, "L13")
# 특기사항 내용 (수행내용/발주처/계약처/계약방식/수금조건/기타)
for c in ["D23", "D24", "D25", "D26", "D27", "D30"]:
    mark(ws, c)
# 보증 조건 선택 + 보증금액
for c in ["P23", "P24", "P25"]:
    mark(ws, c)
for c in ["S23", "S24", "S25"]:
    mark(ws, c)
# 예비비 예상액 (PM 판단값 — 소스 자료 근거 없음, 수동 입력 필요)
mark_orange(ws, "L27")

wb.save(DST)
print(f"템플릿 저장 완료: {DST}")

# ══════════════════════════════════════════════════════
# 자동 검증 1: 초록색 셀 중 수식 셀 → 즉시 오류
# ══════════════════════════════════════════════════════
print("\n[검증 1] 초록색 셀 중 수식 셀 확인...")
wb_chk = openpyxl.load_workbook(DST, data_only=False)
errors = []
for sname in wb_chk.sheetnames:
    ws_c = wb_chk[sname]
    for row in ws_c.iter_rows():
        for cell in row:
            fill = cell.fill
            is_green = fill and fill.fgColor and fill.fgColor.rgb in GREEN_RGB_VARIANTS
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            if is_green and is_formula:
                errors.append(f"  ❌ [{sname}] {cell.coordinate}: {cell.value[:80]}")

if errors:
    for e in errors:
        print(e)
else:
    print("  ✅ 수식 셀 없음")

# ══════════════════════════════════════════════════════
# 자동 검증 2: 실데이터 셀 중 초록색 미표시 확인
# 레이블 판단: C열(비목명), B열, J열(안내문), M열(작성기준),
#             P열(안내문), T열(설명), 행1~5(헤더), #REF! 값
# ══════════════════════════════════════════════════════
print("\n[검증 2] 실데이터 셀 누락 확인...")

wb_src = openpyxl.load_workbook(SRC, data_only=False)
wb_dst = openpyxl.load_workbook(DST, data_only=False)

# 레이블 열: B,C,J,M,P,T + H(수식참조헤더),I(수식),N(안내)
LABEL_COLS = {2, 3, 8, 9, 10, 13, 14, 16, 18, 20, 23, 24, 25}
LABEL_ROWS = {1, 2, 3, 4, 5}

# 공통2: D열 레이블 행 (비목명/섹션명)
COMMON2_D_LABEL_ROWS = {28,29,30,31,32,33,34,35,36,37,38,39,40,41,43,
                         44,45,46,47,48,49,51,52,53,56,57,58,59,61,62}
# 산출내역서: D열 고정 텍스트 행 (보험료 요율 설명 — AI가 수정하지 않음)
SANCH_D_FIXED_ROWS = {26, 27, 28, 29, 30}
# 5.4수수료: 안내문구 행
FEE_LABEL_ROWS = {20, 21, 22, 23, 24, 25}
# 갑지: 고정 레이블 셀
GAPJI_LABEL = {'D38'}  # 회사명 고정값

target = ['공통2', '5.4 수수료', '5. 산출내역서', '0. 집행(갑지)']
missing = []
for sname in target:
    ws_s = wb_src[sname]
    ws_d = wb_dst[sname]
    for row in ws_s.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and (v.strip() == '' or v.startswith('=')):
                continue
            if str(v) == '#REF!':
                continue
            # 레이블 열/행 제외
            if cell.column in LABEL_COLS:
                continue
            if cell.row in LABEL_ROWS:
                continue
            # 시트별 추가 레이블 제외
            if sname == '공통2' and cell.column == 4 and cell.row in COMMON2_D_LABEL_ROWS:
                continue
            if sname == '5.4 수수료' and cell.row in FEE_LABEL_ROWS:
                continue
            # 산출내역서 D7은 안내문구, D26~D30은 보험료 요율 고정 레이블
            if sname == '5. 산출내역서' and cell.coordinate == 'D7':
                continue
            if sname == '5. 산출내역서' and cell.column == 4 and cell.row in SANCH_D_FIXED_ROWS:
                continue
            if sname == '0. 집행(갑지)' and cell.coordinate in GAPJI_LABEL:
                continue
            # 순수 헤더 문자열 제외 (금 액, %, 수 량, 단 가 등)
            if isinstance(v, str) and v.strip() in {
                '금 액', '금액', '%', '수 량', '단 가', '단가', '수량',
                'M/M', '당초', '변경', '증감', '소계', '합계', '소 계', '합     계',
                '(단위 : 원)', '【단위 : 원】', '[단위:천원,%]',
                '■ 재료비', '■ 노무비', '■ 외주비', '■ 경  비', ' 합      계',
                '용역 수수료', '유지보수 수수료', '기타 수수료',
                '보험료(노무)', '   보험료(공사)', '소 계',
                '집행계획(A)', '당기계획(C)', '차기이월(D)', '이월소계(E=C+D)',
                '계  약', '집행계획', '집행계획(A)',
                '수주손익', '계약금액', '집행계획', '정산 누계', '이월 집행 계획', '총계(F=B+E)',
                '당기 계획( C )', '당기 이후(D)', '이월 소계',
                '매 출 액', '재 료 비', '노 무 비', '외 주 비', '경    비', '합    계',
                '영업이익 A', '구  분', '조     건', '보증요율', '보증 금액',
                '선 급 금', '계약이행', '하자보수', '수     금', '원가산출 기준 등',
                '계약물량', '집행물량', '자재단가', '인 건 비',
                '매출', '영업이익 A', '예비비 예상액', '영업이익 B', '(%)',
                '4.예산집행계획', '5. 특기사항', '* 프로젝트 개요', '  *  특기 사항 : ',
                '집행예산 산출내역서', '비목/계정별', '산   출   내   역 ',
                '계약 금액', '집행 금액', '정산 금액(A)', '당기 계획(B)', '차기 이월©', '누계(A+B+C)',
                'GS네오텍                                                       ',
                '     GP-4050-01(1)', 'GP-4050-06(1)',
            }:
                continue
            # 차수 헤더 (최초/1차/2차... 10차)
            if isinstance(v, str) and v in {'최초','1차','2차','3차','4차','5차','6차','7차','8차','9차','10차'}:
                continue
            dst_cell = ws_d[cell.coordinate]
            fill = dst_cell.fill
            rgb = fill.fgColor.rgb if fill and fill.fgColor else ""
            if rgb not in ALL_MARKED:
                missing.append(f"  [{sname}] {cell.coordinate} (col={cell.column},row={cell.row}): {repr(str(v)[:60])}")

if missing:
    print(f"  누락 {len(missing)}개:")
    for m in missing:
        print(m)
else:
    print("  ✅ 누락 없음")
